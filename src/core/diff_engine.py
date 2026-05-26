from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, date, time
from enum import Enum
from typing import Any, Iterable

import openpyxl
from openpyxl.utils import get_column_letter


class DiffStatus(str, Enum):
    EQUAL = "equal"
    MODIFIED = "modified"
    ONLY_LEFT = "only_left"
    ONLY_RIGHT = "only_right"


class SheetPresence(str, Enum):
    BOTH = "both"
    ONLY_LEFT = "only_left"
    ONLY_RIGHT = "only_right"


class AlignmentMode(str, Enum):
    POSITIONAL = "positional"
    KEY_BASED = "key_based"


def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def values_equal(a: Any, b: Any) -> bool:
    a_f = is_formula(a)
    b_f = is_formula(b)
    if a_f != b_f:
        return False
    if a_f:
        return a == b
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(a, bool) and not isinstance(b, bool):
        try:
            return float(a) == float(b)
        except (TypeError, ValueError):
            return False
    return a == b


def is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v == "")


def cell_status(left: Any, right: Any) -> DiffStatus:
    l_empty = is_empty(left)
    r_empty = is_empty(right)
    if l_empty and r_empty:
        return DiffStatus.EQUAL
    if l_empty:
        return DiffStatus.ONLY_RIGHT
    if r_empty:
        return DiffStatus.ONLY_LEFT
    if values_equal(left, right):
        return DiffStatus.EQUAL
    return DiffStatus.MODIFIED


def format_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "VERDADERO" if v else "FALSO"
    if isinstance(v, datetime):
        if v.hour or v.minute or v.second:
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:g}"
    return str(v)


@dataclass
class SheetDiff:
    name: str
    presence: SheetPresence
    # diffs keyed by (visual_row, col). For POSITIONAL mode, visual_row == Excel row.
    # For KEY_BASED mode, visual_row is the index into aligned_rows (1-indexed).
    diffs: dict[tuple[int, int], DiffStatus] = field(default_factory=dict)
    max_row: int = 1
    max_col: int = 1
    alignment_mode: AlignmentMode = AlignmentMode.POSITIONAL
    key_column: int | None = None
    header_row: int = 1
    # aligned_rows[i] = (excel_row_left | None, excel_row_right | None) for visual row i+1
    aligned_rows: list[tuple[int | None, int | None]] = field(default_factory=list)
    # Cached source bounds for re-alignment without re-scanning the workbook
    src_max_row_left: int = 0
    src_max_row_right: int = 0
    src_max_col_left: int = 0
    src_max_col_right: int = 0

    @property
    def diff_count(self) -> int:
        return len(self.diffs)

    def excel_rows_at(self, visual_row: int) -> tuple[int | None, int | None]:
        """Return the (excel_left, excel_right) pair for a 1-indexed visual row."""
        idx = visual_row - 1
        if 0 <= idx < len(self.aligned_rows):
            return self.aligned_rows[idx]
        # Fallback for positional mode without explicit aligned_rows
        return (visual_row, visual_row)


@dataclass
class WorkbookDiff:
    left_path: str
    right_path: str
    sheets: dict[str, SheetDiff] = field(default_factory=dict)

    @property
    def total_diffs(self) -> int:
        return sum(s.diff_count for s in self.sheets.values())


def _sheet_bounds(ws) -> tuple[int, int]:
    """Real content bounds: highest row/col with a non-empty value.

    Some Excel files inflate max_row/max_column with phantom formatting,
    which would otherwise force us to render millions of empty cells.
    """
    reported_r = max(ws.max_row or 0, 0)
    reported_c = max(ws.max_column or 0, 0)
    if reported_r == 0 or reported_c == 0:
        return 1, 1
    max_r = 0
    max_c = 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=reported_r, min_col=1, max_col=reported_c, values_only=True),
        start=1,
    ):
        for col_idx, val in enumerate(row, start=1):
            if val is not None and val != "":
                if row_idx > max_r:
                    max_r = row_idx
                if col_idx > max_c:
                    max_c = col_idx
    return max(max_r, 1), max(max_c, 1)


def build_positional_alignment(mr_l: int, mr_r: int) -> list[tuple[int | None, int | None]]:
    n = max(mr_l, mr_r)
    out: list[tuple[int | None, int | None]] = []
    for r in range(1, n + 1):
        lr = r if r <= mr_l else None
        rr = r if r <= mr_r else None
        out.append((lr, rr))
    return out


def build_key_alignment(
    ws_l,
    ws_r,
    mr_l: int,
    mr_r: int,
    key_col: int,
    header_row: int = 1,
) -> list[tuple[int | None, int | None]]:
    """Pair rows by matching values in `key_col`.

    Header row stays paired (header_row, header_row). Data rows from `ws_l` are
    walked in order and paired with the same key in `ws_r` if found.
    Leftover right rows go at the bottom.
    """
    aligned: list[tuple[int | None, int | None]] = []
    if header_row >= 1:
        aligned.append((header_row, header_row))

    right_by_key: dict[Any, int] = {}
    for r in range(header_row + 1, mr_r + 1):
        k = ws_r.cell(r, key_col).value
        if k is None:
            continue
        if k not in right_by_key:
            right_by_key[k] = r

    used_right: set[int] = {header_row} if header_row >= 1 else set()
    for r in range(header_row + 1, mr_l + 1):
        k = ws_l.cell(r, key_col).value
        rr = right_by_key.get(k) if k is not None else None
        if rr is not None and rr not in used_right:
            aligned.append((r, rr))
            used_right.add(rr)
        else:
            aligned.append((r, None))

    for r in range(header_row + 1, mr_r + 1):
        if r not in used_right:
            aligned.append((None, r))

    return aligned


def detect_key_column(
    ws,
    mr: int,
    mc: int,
    header_row: int = 1,
    scan_limit: int = 12,
) -> int | None:
    """Find the leftmost column whose non-empty values are all unique.

    Treats it as a key if at least half of the data rows have a non-empty value
    and there are no duplicates among the non-empty values.
    """
    n_data = mr - header_row
    if n_data < 2:
        return None
    end_col = min(mc, scan_limit)
    for c in range(1, end_col + 1):
        seen: set | None = set()
        n_nonempty = 0
        for r in range(header_row + 1, mr + 1):
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            if v in seen:
                seen = None
                break
            seen.add(v)
            n_nonempty += 1
        if seen is not None and n_nonempty >= max(2, n_data // 2):
            return c
    return None


def diffs_from_alignment(
    ws_l,
    ws_r,
    aligned_rows: list[tuple[int | None, int | None]],
    max_col: int,
) -> dict[tuple[int, int], DiffStatus]:
    """Compute the per-cell diff status keyed by (visual_row, col)."""
    diffs: dict[tuple[int, int], DiffStatus] = {}
    for v_idx, (lr, rr) in enumerate(aligned_rows, start=1):
        for c in range(1, max_col + 1):
            lv = ws_l.cell(lr, c).value if lr is not None else None
            rv = ws_r.cell(rr, c).value if rr is not None else None
            st = cell_status(lv, rv)
            if st != DiffStatus.EQUAL:
                diffs[(v_idx, c)] = st
    return diffs


def _diff_sheet_both(name: str, ws_l, ws_r) -> SheetDiff:
    mr_l, mc_l = _sheet_bounds(ws_l)
    mr_r, mc_r = _sheet_bounds(ws_r)
    max_col = max(mc_l, mc_r)
    aligned = build_positional_alignment(mr_l, mr_r)
    diffs = diffs_from_alignment(ws_l, ws_r, aligned, max_col)
    return SheetDiff(
        name=name,
        presence=SheetPresence.BOTH,
        diffs=diffs,
        max_row=len(aligned),
        max_col=max_col,
        alignment_mode=AlignmentMode.POSITIONAL,
        aligned_rows=aligned,
        src_max_row_left=mr_l,
        src_max_row_right=mr_r,
        src_max_col_left=mc_l,
        src_max_col_right=mc_r,
    )


def realign_sheet(
    sd: SheetDiff,
    ws_l,
    ws_r,
    mode: AlignmentMode,
    key_col: int | None = None,
    header_row: int = 1,
) -> None:
    """Recompute the alignment and diffs for a sheet in place."""
    mr_l = sd.src_max_row_left or _sheet_bounds(ws_l)[0]
    mr_r = sd.src_max_row_right or _sheet_bounds(ws_r)[0]
    max_col = max(sd.src_max_col_left, sd.src_max_col_right) or max(
        _sheet_bounds(ws_l)[1], _sheet_bounds(ws_r)[1]
    )
    if mode == AlignmentMode.POSITIONAL:
        aligned = build_positional_alignment(mr_l, mr_r)
        sd.key_column = None
    else:
        if key_col is None:
            key_col = detect_key_column(ws_l, mr_l, max_col, header_row=header_row)
            if key_col is None:
                key_col = detect_key_column(ws_r, mr_r, max_col, header_row=header_row)
            if key_col is None:
                key_col = 1
        aligned = build_key_alignment(ws_l, ws_r, mr_l, mr_r, key_col, header_row=header_row)
        sd.key_column = key_col
    sd.aligned_rows = aligned
    sd.alignment_mode = mode
    sd.max_row = len(aligned)
    sd.max_col = max_col
    sd.header_row = header_row
    sd.diffs = diffs_from_alignment(ws_l, ws_r, aligned, max_col)


def _diff_sheet_one_side(name: str, ws, side: str) -> SheetDiff:
    mr, mc = _sheet_bounds(ws)
    diffs: dict[tuple[int, int], DiffStatus] = {}
    status = DiffStatus.ONLY_LEFT if side == "left" else DiffStatus.ONLY_RIGHT
    aligned: list[tuple[int | None, int | None]] = []
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if not is_empty(v):
                diffs[(r, c)] = status
        aligned.append((r, None) if side == "left" else (None, r))
    presence = SheetPresence.ONLY_LEFT if side == "left" else SheetPresence.ONLY_RIGHT
    return SheetDiff(
        name=name,
        presence=presence,
        diffs=diffs,
        max_row=mr,
        max_col=mc,
        aligned_rows=aligned,
        src_max_row_left=mr if side == "left" else 0,
        src_max_row_right=mr if side == "right" else 0,
        src_max_col_left=mc if side == "left" else 0,
        src_max_col_right=mc if side == "right" else 0,
    )


def compute_workbook_diff(wb_left, wb_right, left_path: str, right_path: str) -> WorkbookDiff:
    left_names = list(wb_left.sheetnames)
    right_names = list(wb_right.sheetnames)
    seen: set[str] = set()
    ordered: list[str] = []
    for n in left_names + right_names:
        if n not in seen:
            ordered.append(n)
            seen.add(n)

    sheets: dict[str, SheetDiff] = {}
    for name in ordered:
        in_l = name in wb_left.sheetnames
        in_r = name in wb_right.sheetnames
        if in_l and in_r:
            sheets[name] = _diff_sheet_both(name, wb_left[name], wb_right[name])
        elif in_l:
            sheets[name] = _diff_sheet_one_side(name, wb_left[name], "left")
        else:
            sheets[name] = _diff_sheet_one_side(name, wb_right[name], "right")
    return WorkbookDiff(left_path=left_path, right_path=right_path, sheets=sheets)


def recompute_cell_status(ws_left, ws_right, row: int, col: int) -> DiffStatus:
    """Re-evaluate the diff status of a single cell after an edit."""
    lv = ws_left.cell(row, col).value if ws_left is not None else None
    rv = ws_right.cell(row, col).value if ws_right is not None else None
    return cell_status(lv, rv)


def cell_coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def sorted_diff_keys(sd: SheetDiff) -> list[tuple[int, int]]:
    return sorted(sd.diffs.keys())
