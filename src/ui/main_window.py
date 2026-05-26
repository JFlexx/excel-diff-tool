from __future__ import annotations

import os
import shutil
import traceback
from pathlib import Path

import openpyxl
from PySide6.QtCore import Qt, QObject, QThread, Signal
from PySide6.QtGui import QAction, QCloseEvent, QFont, QIcon, QKeySequence
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QStatusBar,
    QTabWidget,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from src.core.diff_engine import (
    SheetDiff,
    SheetPresence,
    WorkbookDiff,
    cell_coord,
    compute_workbook_diff,
)
from src.ui.diff_view import DiffView


class TabPanel(QWidget):
    """Tab container that materializes its DiffView on first show.

    Avoids removeTab/insertTab swaps which can re-enter currentChanged.
    """
    left_modified = Signal()
    right_modified = Signal()
    selection_changed = Signal(int, int, str)

    def __init__(
        self,
        sheet_diff: SheetDiff,
        wb_left,
        wb_right,
        left_name: str,
        right_name: str,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.sheet_diff = sheet_diff
        self.wb_left = wb_left
        self.wb_right = wb_right
        self.left_name = left_name
        self.right_name = right_name
        self.view: DiffView | None = None
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self._placeholder = QLabel("Cargando hoja…")
        self._placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setStyleSheet("color:#888; padding:40px;")
        layout.addWidget(self._placeholder)

    def materialize(self) -> DiffView:
        if self.view is not None:
            return self.view
        name = self.sheet_diff.name
        ws_l = self.wb_left[name] if self.wb_left is not None and name in self.wb_left.sheetnames else None
        ws_r = self.wb_right[name] if self.wb_right is not None and name in self.wb_right.sheetnames else None
        self.view = DiffView(
            ws_l, ws_r, self.sheet_diff,
            left_name=self.left_name, right_name=self.right_name,
            parent=self,
        )
        self.view.left_modified.connect(self.left_modified)
        self.view.right_modified.connect(self.right_modified)
        self.view.selection_changed.connect(self.selection_changed)
        layout = self.layout()
        layout.removeWidget(self._placeholder)
        self._placeholder.deleteLater()
        self._placeholder = None
        layout.addWidget(self.view)
        return self.view


class DiffWorker(QObject):
    finished = Signal(object, object, object)  # wb_left, wb_right, diff
    failed = Signal(str)

    def __init__(self, left: str, right: str) -> None:
        super().__init__()
        self.left = left
        self.right = right

    def run(self) -> None:
        try:
            wb_l = openpyxl.load_workbook(self.left, data_only=False)
            wb_r = openpyxl.load_workbook(self.right, data_only=False)
            diff = compute_workbook_diff(wb_l, wb_r, self.left, self.right)
            self.finished.emit(wb_l, wb_r, diff)
        except Exception:
            self.failed.emit(traceback.format_exc())


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Excel Diff Tool")
        self.resize(1500, 900)

        self.wb_left = None
        self.wb_right = None
        self.diff: WorkbookDiff | None = None
        self.left_dirty = False
        self.right_dirty = False
        self.left_name = "A"
        self.right_name = "B"
        self._tab_sheet_names: list[str] = []
        self._thread: QThread | None = None
        self._worker: DiffWorker | None = None
        self._progress: QProgressDialog | None = None

        self._build_ui()
        self._refresh_action_state()

    # ---------- UI ----------

    def _build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        root.addWidget(self._build_files_bar())
        root.addWidget(self._build_action_bar())

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self._on_tab_changed)
        root.addWidget(self.tabs, 1)

        root.addWidget(self._build_save_bar())

        self.setStatusBar(QStatusBar())
        self.statusBar().showMessage("Selecciona los dos ficheros Excel y pulsa «Comparar».")

    def _build_files_bar(self) -> QWidget:
        box = QFrame()
        box.setFrameShape(QFrame.Shape.StyledPanel)
        layout = QHBoxLayout(box)
        layout.setContentsMargins(8, 6, 8, 6)

        layout.addWidget(QLabel("Fichero A:"))
        self.left_path = QLineEdit()
        self.left_path.setPlaceholderText("Primer Excel a comparar…")
        layout.addWidget(self.left_path, 1)
        btn_l = QPushButton("Examinar…")
        btn_l.clicked.connect(lambda: self._pick_file(self.left_path))
        layout.addWidget(btn_l)

        layout.addSpacing(12)
        layout.addWidget(QLabel("Fichero B:"))
        self.right_path = QLineEdit()
        self.right_path.setPlaceholderText("Segundo Excel a comparar…")
        layout.addWidget(self.right_path, 1)
        btn_r = QPushButton("Examinar…")
        btn_r.clicked.connect(lambda: self._pick_file(self.right_path))
        layout.addWidget(btn_r)

        layout.addSpacing(12)
        self.btn_compare = QPushButton("Comparar")
        self.btn_compare.setMinimumWidth(110)
        self.btn_compare.clicked.connect(self._start_compare)
        layout.addWidget(self.btn_compare)
        return box

    def _build_action_bar(self) -> QWidget:
        box = QFrame()
        box.setFrameShape(QFrame.Shape.StyledPanel)
        layout = QHBoxLayout(box)
        layout.setContentsMargins(8, 4, 8, 4)

        bold = QFont()
        bold.setBold(True)

        self.btn_copy_to_left = QPushButton("◀  Tomar de B")
        self.btn_copy_to_left.setToolTip("Copia el valor de la celda seleccionada de B a A. (Alt+Izquierda)")
        self.btn_copy_to_left.setMinimumHeight(34)
        self.btn_copy_to_left.setFont(bold)
        self.btn_copy_to_left.clicked.connect(self._copy_to_left)
        layout.addWidget(self.btn_copy_to_left)

        self.btn_copy_to_right = QPushButton("Tomar de A  ▶")
        self.btn_copy_to_right.setToolTip("Copia el valor de la celda seleccionada de A a B. (Alt+Derecha)")
        self.btn_copy_to_right.setMinimumHeight(34)
        self.btn_copy_to_right.setFont(bold)
        self.btn_copy_to_right.clicked.connect(self._copy_to_right)
        layout.addWidget(self.btn_copy_to_right)

        layout.addSpacing(16)

        self.btn_prev = QPushButton("‹ Anterior diferencia")
        self.btn_prev.setToolTip("Saltar a la diferencia anterior. (Shift+F8)")
        self.btn_prev.clicked.connect(self._prev_diff)
        layout.addWidget(self.btn_prev)

        self.btn_next = QPushButton("Siguiente diferencia ›")
        self.btn_next.setToolTip("Saltar a la siguiente diferencia. (F8)")
        self.btn_next.clicked.connect(self._next_diff)
        layout.addWidget(self.btn_next)

        layout.addStretch(1)

        self.cell_label = QLabel("—")
        self.cell_label.setStyleSheet("padding:0 8px;")
        layout.addWidget(self.cell_label)

        return box

    def _build_save_bar(self) -> QWidget:
        box = QFrame()
        box.setFrameShape(QFrame.Shape.StyledPanel)
        layout = QHBoxLayout(box)
        layout.setContentsMargins(8, 6, 8, 6)

        self.stats_label = QLabel("Sin comparación.")
        layout.addWidget(self.stats_label, 1)

        self.btn_save_left = QPushButton("Guardar A")
        self.btn_save_left.setMinimumWidth(200)
        self.btn_save_left.clicked.connect(lambda: self._save_side("left"))
        layout.addWidget(self.btn_save_left)

        self.btn_save_right = QPushButton("Guardar B")
        self.btn_save_right.setMinimumWidth(200)
        self.btn_save_right.clicked.connect(lambda: self._save_side("right"))
        layout.addWidget(self.btn_save_right)

        return box

    # ---------- File picker ----------

    def _pick_file(self, target: QLineEdit) -> None:
        start = target.text() or str(Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "Selecciona un Excel", start, "Excel (*.xlsx *.xlsm)"
        )
        if path:
            target.setText(path)

    # ---------- Compare ----------

    def _start_compare(self) -> None:
        if not self._confirm_discard_dirty():
            return
        left = self.left_path.text().strip()
        right = self.right_path.text().strip()
        if not left or not right:
            QMessageBox.warning(self, "Faltan ficheros", "Selecciona ambos ficheros antes de comparar.")
            return
        if not os.path.isfile(left) or not os.path.isfile(right):
            QMessageBox.warning(self, "Ruta inválida", "Una de las rutas no existe.")
            return

        self.btn_compare.setEnabled(False)
        self._progress = QProgressDialog("Comparando ficheros…", None, 0, 0, self)
        self._progress.setWindowModality(Qt.WindowModality.WindowModal)
        self._progress.setCancelButton(None)
        self._progress.setMinimumDuration(0)
        self._progress.show()

        self._thread = QThread(self)
        self._worker = DiffWorker(left, right)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished.connect(self._on_diff_ready)
        self._worker.failed.connect(self._on_diff_failed)
        self._worker.finished.connect(self._thread.quit)
        self._worker.failed.connect(self._thread.quit)
        self._thread.finished.connect(self._cleanup_thread)
        self._thread.start()

    def _cleanup_thread(self) -> None:
        if self._progress:
            self._progress.close()
            self._progress = None
        if self._thread:
            self._thread.deleteLater()
            self._thread = None
        if self._worker:
            self._worker.deleteLater()
            self._worker = None
        self.btn_compare.setEnabled(True)

    def _on_diff_failed(self, message: str) -> None:
        QMessageBox.critical(self, "Error al comparar", message)
        self.statusBar().showMessage("Comparación fallida.")

    def _on_diff_ready(self, wb_left, wb_right, diff: WorkbookDiff) -> None:
        self.wb_left = wb_left
        self.wb_right = wb_right
        self.diff = diff
        self.left_dirty = False
        self.right_dirty = False
        self.left_name, self.right_name = self._derive_display_names(diff.left_path, diff.right_path)
        self._update_named_actions()
        self._populate_tabs()
        self._refresh_stats()
        self._refresh_action_state()
        msg = f"{diff.total_diffs} celdas con diferencias."
        if diff.total_diffs == 0:
            msg = "Los ficheros son idénticos."
        self.statusBar().showMessage(msg)

    @staticmethod
    def _derive_display_names(left_path: str, right_path: str) -> tuple[str, str]:
        l = os.path.basename(left_path)
        r = os.path.basename(right_path)
        if l == r:
            # Same basename in different folders: disambiguate by parent folder.
            l = f"{os.path.basename(os.path.dirname(left_path))}/{l}"
            r = f"{os.path.basename(os.path.dirname(right_path))}/{r}"
        return l, r

    @staticmethod
    def _short(name: str, limit: int = 28) -> str:
        if len(name) <= limit:
            return name
        return name[: limit - 1] + "…"

    def _update_named_actions(self) -> None:
        a = self._short(self.left_name)
        b = self._short(self.right_name)
        self.btn_copy_to_left.setText(f"◀  Tomar de {b}")
        self.btn_copy_to_left.setToolTip(f"Copia el valor de la celda seleccionada de «{self.right_name}» a «{self.left_name}». (Alt+Izquierda)")
        self.btn_copy_to_right.setText(f"Tomar de {a}  ▶")
        self.btn_copy_to_right.setToolTip(f"Copia el valor de la celda seleccionada de «{self.left_name}» a «{self.right_name}». (Alt+Derecha)")
        self.btn_save_left.setText(f"Guardar {a}")
        self.btn_save_right.setText(f"Guardar {b}")

    # ---------- Tabs ----------

    def _populate_tabs(self) -> None:
        self.tabs.blockSignals(True)
        self.tabs.clear()
        self._tab_sheet_names = []
        if not self.diff:
            self.tabs.blockSignals(False)
            return

        for name, sd in self.diff.sheets.items():
            self._tab_sheet_names.append(name)
            panel = TabPanel(sd, self.wb_left, self.wb_right, self.left_name, self.right_name)
            panel.left_modified.connect(lambda p=panel: self._on_panel_modified(p, "left"))
            panel.right_modified.connect(lambda p=panel: self._on_panel_modified(p, "right"))
            panel.selection_changed.connect(self._on_selection_changed)
            self.tabs.addTab(panel, self._format_tab_label(sd))
        self.tabs.blockSignals(False)
        if self.tabs.count() > 0:
            self.tabs.setCurrentIndex(0)
            self._on_tab_changed(0)

    def _format_tab_label(self, sd: SheetDiff) -> str:
        if sd.presence == SheetPresence.ONLY_LEFT:
            return f"{sd.name}  ◐ solo en A"
        if sd.presence == SheetPresence.ONLY_RIGHT:
            return f"{sd.name}  ◑ solo en B"
        if sd.diff_count == 0:
            return f"{sd.name}  ✓"
        return f"{sd.name}  ({sd.diff_count})"

    def _refresh_tab_label(self, idx: int) -> None:
        if not self.diff or idx < 0 or idx >= len(self._tab_sheet_names):
            return
        name = self._tab_sheet_names[idx]
        self.tabs.setTabText(idx, self._format_tab_label(self.diff.sheets[name]))

    def _on_tab_changed(self, idx: int) -> None:
        if idx < 0:
            return
        panel = self.tabs.widget(idx)
        if isinstance(panel, TabPanel) and panel.view is None:
            # Materialize inside a try so a failure doesn't crash the app silently
            try:
                panel.materialize()
            except Exception:
                QMessageBox.critical(self, "Error al cargar la hoja", traceback.format_exc())
        self._refresh_action_state()

    def _current_panel(self) -> TabPanel | None:
        idx = self.tabs.currentIndex()
        if idx < 0:
            return None
        w = self.tabs.widget(idx)
        return w if isinstance(w, TabPanel) else None

    def _current_view(self) -> DiffView | None:
        panel = self._current_panel()
        return panel.view if panel else None

    def _on_panel_modified(self, panel: TabPanel, side: str) -> None:
        idx = self.tabs.indexOf(panel)
        if idx >= 0:
            self._on_side_modified(idx, side)

    # ---------- Actions ----------

    def _copy_to_left(self) -> None:
        v = self._current_view()
        if v:
            v.copy_right_to_left()

    def _copy_to_right(self) -> None:
        v = self._current_view()
        if v:
            v.copy_left_to_right()

    def _next_diff(self) -> None:
        v = self._current_view()
        if v and not v.goto_next_diff():
            self.statusBar().showMessage("No hay más diferencias en esta hoja.", 3000)

    def _prev_diff(self) -> None:
        v = self._current_view()
        if v and not v.goto_prev_diff():
            self.statusBar().showMessage("No hay más diferencias en esta hoja.", 3000)

    def _on_side_modified(self, tab_idx: int, side: str) -> None:
        if side == "left":
            self.left_dirty = True
        else:
            self.right_dirty = True
        self._refresh_tab_label(tab_idx)
        self._refresh_stats()
        self._refresh_action_state()

    def _on_selection_changed(self, row: int, col: int, status: str) -> None:
        label_map = {
            "modified": "Modificada",
            "only_left": f"Solo en {self._short(self.left_name)}",
            "only_right": f"Solo en {self._short(self.right_name)}",
            "equal": "Igual",
        }
        self.cell_label.setText(f"Celda {cell_coord(row, col)}  ·  {label_map.get(status, '')}")

    # ---------- Stats / state ----------

    def _refresh_stats(self) -> None:
        if not self.diff:
            self.stats_label.setText("Sin comparación.")
            return
        total = sum(sd.diff_count for sd in self.diff.sheets.values())
        bits = [f"{total} celdas con diferencias"]
        if self.left_dirty:
            bits.append(f"{self._short(self.left_name)} modificado ●")
        if self.right_dirty:
            bits.append(f"{self._short(self.right_name)} modificado ●")
        self.stats_label.setText("  ·  ".join(bits))

    def _refresh_action_state(self) -> None:
        v = self._current_view()
        has_view = v is not None
        both_sides = has_view and v.sheet_diff.presence == SheetPresence.BOTH
        self.btn_copy_to_left.setEnabled(both_sides)
        self.btn_copy_to_right.setEnabled(both_sides)
        self.btn_next.setEnabled(has_view)
        self.btn_prev.setEnabled(has_view)
        self.btn_save_left.setEnabled(self.wb_left is not None)
        self.btn_save_right.setEnabled(self.wb_right is not None)
        a = self._short(self.left_name)
        b = self._short(self.right_name)
        self.btn_save_left.setText(f"Guardar {a} ●" if self.left_dirty else f"Guardar {a}")
        self.btn_save_right.setText(f"Guardar {b} ●" if self.right_dirty else f"Guardar {b}")
        title = "Excel Diff Tool"
        if self.left_dirty or self.right_dirty:
            title += "  *"
        self.setWindowTitle(title)

    # ---------- Save ----------

    def _save_side(self, side: str) -> None:
        if not self.diff:
            return
        if side == "left":
            path = self.diff.left_path
            wb = self.wb_left
            dirty = self.left_dirty
        else:
            path = self.diff.right_path
            wb = self.wb_right
            dirty = self.right_dirty
        if not dirty:
            QMessageBox.information(self, "Nada que guardar", f"El fichero {os.path.basename(path)} no tiene cambios.")
            return
        ans = QMessageBox.question(
            self, "Sobreescribir fichero",
            f"Vas a sobrescribir:\n{path}\n\n"
            "Se hará una copia de seguridad junto al fichero (.bak) la primera vez.\n\n¿Continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Yes,
        )
        if ans != QMessageBox.StandardButton.Yes:
            return
        try:
            self._ensure_backup(path)
            wb.save(path)
        except PermissionError:
            QMessageBox.critical(
                self, "Fichero en uso",
                f"No se puede escribir en {path}.\n\n"
                "¿Está abierto en Excel? Ciérralo e inténtalo de nuevo."
            )
            return
        except Exception:
            QMessageBox.critical(self, "Error al guardar", traceback.format_exc())
            return

        if side == "left":
            self.left_dirty = False
        else:
            self.right_dirty = False
        self._refresh_stats()
        self._refresh_action_state()
        self.statusBar().showMessage(f"Guardado: {path}", 5000)

    def _ensure_backup(self, path: str) -> None:
        bak = f"{path}.bak"
        if not os.path.exists(bak):
            shutil.copy2(path, bak)

    # ---------- Close confirmation ----------

    def _confirm_discard_dirty(self) -> bool:
        if not (self.left_dirty or self.right_dirty):
            return True
        ans = QMessageBox.question(
            self, "Cambios sin guardar",
            "Tienes cambios sin guardar. ¿Descartarlos y continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        return ans == QMessageBox.StandardButton.Yes

    def closeEvent(self, event: QCloseEvent) -> None:
        if self._confirm_discard_dirty():
            event.accept()
        else:
            event.ignore()
