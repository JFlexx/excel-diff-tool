from __future__ import annotations

from typing import Any

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QBrush, QColor, QFont, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from openpyxl.utils import get_column_letter

from src.core.diff_engine import (
    AlignmentMode,
    DiffStatus,
    SheetDiff,
    SheetPresence,
    cell_coord,
    format_value,
    is_formula,
    realign_sheet,
    recompute_cell_status,
)
from src.ui.styles import cell_color, detect_theme


MAX_DISPLAY_CELLS = 500_000  # safety cap: above this we warn the user


def _display_text(value: Any) -> str:
    if is_formula(value):
        return value
    return format_value(value)


def _parse_text(text: str) -> Any:
    if text == "":
        return None
    s = text.strip()
    if s.startswith("="):
        return s
    # int
    try:
        return int(s)
    except ValueError:
        pass
    # float (accept comma decimal)
    try:
        return float(s.replace(",", "."))
    except ValueError:
        pass
    return text  # keep original (preserves leading/trailing spaces if intentional)


class DiffView(QWidget):
    left_modified = Signal()
    right_modified = Signal()
    selection_changed = Signal(int, int, str, object, object)  # row, col, status, left_value, right_value
    alignment_changed = Signal()  # emitted after re-alignment so the host can refresh labels

    def __init__(
        self,
        ws_left,
        ws_right,
        sheet_diff: SheetDiff,
        left_name: str = "A",
        right_name: str = "B",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.ws_left = ws_left      # may be None if sheet absent on left
        self.ws_right = ws_right    # may be None if sheet absent on right
        self.sheet_diff = sheet_diff
        self.left_name = left_name
        self.right_name = right_name
        self._syncing_scroll = False
        self._syncing_sel = False
        self._programmatic = False
        self._cap_banner: QWidget | None = None
        self._build_ui()
        self._populate()
        self._wire_sync()
        self._install_shortcuts()

    def refresh(self) -> None:
        """Re-render after the underlying SheetDiff has been re-aligned."""
        self._populate()

    # ---------- UI ----------

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        if self.sheet_diff.presence == SheetPresence.BOTH:
            layout.addWidget(self._build_alignment_bar())

        if self.sheet_diff.presence != SheetPresence.BOTH:
            theme = detect_theme()
            banner = QFrame()
            banner.setFrameShape(QFrame.Shape.StyledPanel)
            banner.setStyleSheet(f"background:{theme.modified.name()}; padding:6px;")
            bl = QHBoxLayout(banner)
            bl.setContentsMargins(8, 4, 8, 4)
            existing = self.left_name if self.sheet_diff.presence == SheetPresence.ONLY_LEFT else self.right_name
            bl.addWidget(QLabel(
                f"Esta hoja solo existe en <b>{existing}</b>. "
                f"La comparación celda a celda no aplica."
            ))
            layout.addWidget(banner)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.left_table = self._make_table(self.left_name)
        self.right_table = self._make_table(self.right_name)

        splitter.addWidget(self._wrap_with_header(self.left_table, self.left_name))
        splitter.addWidget(self._wrap_with_header(self.right_table, self.right_name))
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)
        layout.addWidget(splitter, 1)

    def _wrap_with_header(self, table: QTableWidget, title: str) -> QWidget:
        wrap = QWidget()
        v = QVBoxLayout(wrap)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(0)
        theme = detect_theme()
        lbl = QLabel(title)
        f = QFont()
        f.setBold(True)
        lbl.setFont(f)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(
            f"background:{theme.header_bg.name()};"
            f"color:{theme.header_fg.name()};"
            f"padding:4px;"
        )
        lbl.setToolTip(title)
        v.addWidget(lbl)
        v.addWidget(table, 1)
        return wrap

    # ---------- Alignment bar ----------

    def _build_alignment_bar(self) -> QWidget:
        bar = QFrame()
        bar.setFrameShape(QFrame.Shape.NoFrame)
        bl = QHBoxLayout(bar)
        bl.setContentsMargins(4, 2, 4, 2)
        bl.setSpacing(6)

        bl.addWidget(QLabel("Alineación:"))
        self.align_combo = QComboBox()
        self.align_combo.addItem("Por posición (fila → fila)", AlignmentMode.POSITIONAL.value)
        self.align_combo.addItem("Por columna clave (emparejar por valor)", AlignmentMode.KEY_BASED.value)
        self.align_combo.currentIndexChanged.connect(self._on_align_choice)
        bl.addWidget(self.align_combo)

        bl.addSpacing(8)
        self.key_label = QLabel("Columna clave:")
        bl.addWidget(self.key_label)
        self.key_combo = QComboBox()
        self.key_combo.setMinimumWidth(220)
        self.key_combo.currentIndexChanged.connect(self._on_key_changed)
        bl.addWidget(self.key_combo)

        self._populate_key_combo()
        self._sync_alignment_controls()

        bl.addStretch(1)
        return bar

    def _populate_key_combo(self) -> None:
        self.key_combo.blockSignals(True)
        self.key_combo.clear()
        cols = max(self.sheet_diff.src_max_col_left, self.sheet_diff.src_max_col_right) or self.sheet_diff.max_col
        header_row = self.sheet_diff.header_row or 1
        for c in range(1, cols + 1):
            letter = get_column_letter(c)
            header_text: str | None = None
            if self.ws_left is not None:
                v = self.ws_left.cell(header_row, c).value
                if v is not None and str(v).strip():
                    header_text = str(v)
            if header_text is None and self.ws_right is not None:
                v = self.ws_right.cell(header_row, c).value
                if v is not None and str(v).strip():
                    header_text = str(v)
            label = f"{letter}: {header_text}" if header_text else letter
            self.key_combo.addItem(label, c)
        self.key_combo.blockSignals(False)

    def _sync_alignment_controls(self) -> None:
        """Reflect the SheetDiff's current alignment mode and key in the combos."""
        is_key = self.sheet_diff.alignment_mode == AlignmentMode.KEY_BASED
        self.align_combo.blockSignals(True)
        self.align_combo.setCurrentIndex(1 if is_key else 0)
        self.align_combo.blockSignals(False)
        self.key_combo.setEnabled(is_key)
        self.key_label.setEnabled(is_key)
        if is_key and self.sheet_diff.key_column is not None:
            idx = self.sheet_diff.key_column - 1
            if 0 <= idx < self.key_combo.count():
                self.key_combo.blockSignals(True)
                self.key_combo.setCurrentIndex(idx)
                self.key_combo.blockSignals(False)

    def _on_align_choice(self, _idx: int) -> None:
        mode = AlignmentMode(self.align_combo.currentData())
        self.key_combo.setEnabled(mode == AlignmentMode.KEY_BASED)
        self.key_label.setEnabled(mode == AlignmentMode.KEY_BASED)
        key_col = self.key_combo.currentData() if mode == AlignmentMode.KEY_BASED else None
        self._apply_alignment(mode, key_col)

    def _on_key_changed(self, _idx: int) -> None:
        mode = AlignmentMode(self.align_combo.currentData())
        if mode != AlignmentMode.KEY_BASED:
            return
        self._apply_alignment(mode, self.key_combo.currentData())

    def _apply_alignment(self, mode: AlignmentMode, key_col: int | None) -> None:
        if self.ws_left is None or self.ws_right is None:
            return
        realign_sheet(self.sheet_diff, self.ws_left, self.ws_right, mode, key_col=key_col)
        self._sync_alignment_controls()
        self._populate()
        self.alignment_changed.emit()

    def _make_table(self, _name: str) -> QTableWidget:
        t = QTableWidget()
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        t.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        t.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        t.setAlternatingRowColors(False)
        t.horizontalHeader().setDefaultSectionSize(120)
        t.verticalHeader().setDefaultSectionSize(22)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        return t

    # ---------- Populate ----------

    def _populate(self) -> None:
        sd = self.sheet_diff
        rows = max(sd.max_row, 1)
        cols = max(sd.max_col, 1)

        if self._cap_banner is not None:
            self._cap_banner.setParent(None)
            self._cap_banner.deleteLater()
            self._cap_banner = None

        if rows * cols > MAX_DISPLAY_CELLS:
            if sd.diffs:
                rows = max(k[0] for k in sd.diffs)
                cols = max(k[1] for k in sd.diffs)
            while rows * cols > MAX_DISPLAY_CELLS and rows > 1:
                rows = max(rows // 2, 1)
            theme = detect_theme()
            cap_banner = QFrame()
            cap_banner.setStyleSheet(f"background:{theme.modified.name()}; padding:6px;")
            cl = QHBoxLayout(cap_banner)
            cl.setContentsMargins(8, 4, 8, 4)
            cl.addWidget(QLabel(
                f"<b>Aviso:</b> esta hoja supera el límite seguro de renderizado. "
                f"Mostrando hasta fila {rows}, columna {cols}. "
                f"Las celdas fuera de este rango siguen en el fichero pero no son editables aquí."
            ))
            self.layout().insertWidget(0, cap_banner)
            self._cap_banner = cap_banner

        col_labels = [get_column_letter(i + 1) for i in range(cols)]
        # Per-side row labels: each side shows its own Excel row number; gap = "—"
        left_row_labels = []
        right_row_labels = []
        for v_idx in range(1, rows + 1):
            lr, rr = sd.excel_rows_at(v_idx) if sd.aligned_rows else (v_idx, v_idx)
            left_row_labels.append(str(lr) if lr is not None else "—")
            right_row_labels.append(str(rr) if rr is not None else "—")

        for tbl, row_labels in (
            (self.left_table, left_row_labels),
            (self.right_table, right_row_labels),
        ):
            self._programmatic = True
            tbl.blockSignals(True)
            tbl.clear()
            tbl.setRowCount(rows)
            tbl.setColumnCount(cols)
            tbl.setHorizontalHeaderLabels(col_labels)
            tbl.setVerticalHeaderLabels(row_labels)
            tbl.blockSignals(False)
            self._programmatic = False

        # Fill cells
        self._programmatic = True
        self.left_table.blockSignals(True)
        self.right_table.blockSignals(True)
        for v_idx in range(1, rows + 1):
            if sd.aligned_rows:
                lr, rr = sd.excel_rows_at(v_idx)
            else:
                lr, rr = v_idx, v_idx
            for c in range(1, cols + 1):
                lv = self.ws_left.cell(lr, c).value if lr is not None and self.ws_left is not None else None
                rv = self.ws_right.cell(rr, c).value if rr is not None and self.ws_right is not None else None
                self._set_item(self.left_table, v_idx, c, lv, editable=lr is not None)
                self._set_item(self.right_table, v_idx, c, rv, editable=rr is not None)
        for (r, c), status in sd.diffs.items():
            self._paint(self.left_table, r, c, cell_color(status.value, "left"))
            self._paint(self.right_table, r, c, cell_color(status.value, "right"))
        self.left_table.blockSignals(False)
        self.right_table.blockSignals(False)
        self._programmatic = False

    def _set_item(self, table: QTableWidget, row: int, col: int, value: Any, editable: bool = True) -> None:
        text = _display_text(value)
        item = QTableWidgetItem(text)
        item.setData(Qt.ItemDataRole.UserRole, value)
        if len(text) > 20:
            item.setToolTip(text)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        table.setItem(row - 1, col - 1, item)

    def _paint(self, table: QTableWidget, row: int, col: int, color: QColor) -> None:
        item = table.item(row - 1, col - 1)
        if item is None:
            item = QTableWidgetItem("")
            table.setItem(row - 1, col - 1, item)
        item.setBackground(QBrush(color))

    # ---------- Sync ----------

    def _wire_sync(self) -> None:
        # Scroll sync (vertical + horizontal)
        l_v = self.left_table.verticalScrollBar()
        r_v = self.right_table.verticalScrollBar()
        l_h = self.left_table.horizontalScrollBar()
        r_h = self.right_table.horizontalScrollBar()
        l_v.valueChanged.connect(lambda v: self._sync_scroll(r_v, v))
        r_v.valueChanged.connect(lambda v: self._sync_scroll(l_v, v))
        l_h.valueChanged.connect(lambda v: self._sync_scroll(r_h, v))
        r_h.valueChanged.connect(lambda v: self._sync_scroll(l_h, v))

        # Selection sync
        self.left_table.currentCellChanged.connect(self._on_left_current)
        self.right_table.currentCellChanged.connect(self._on_right_current)

        # Inline edit detection
        self.left_table.itemChanged.connect(self._on_left_edited)
        self.right_table.itemChanged.connect(self._on_right_edited)

    def _sync_scroll(self, target_sb, value: int) -> None:
        if self._syncing_scroll:
            return
        self._syncing_scroll = True
        target_sb.setValue(value)
        self._syncing_scroll = False

    def _on_left_current(self, r: int, c: int, _pr: int, _pc: int) -> None:
        if not self._syncing_sel:
            self._syncing_sel = True
            self.right_table.setCurrentCell(r, c)
            self._syncing_sel = False
        self._emit_selection(r, c)

    def _on_right_current(self, r: int, c: int, _pr: int, _pc: int) -> None:
        if not self._syncing_sel:
            self._syncing_sel = True
            self.left_table.setCurrentCell(r, c)
            self._syncing_sel = False
        self._emit_selection(r, c)

    def _emit_selection(self, r: int, c: int) -> None:
        if r < 0 or c < 0:
            return
        v_row = r + 1
        col = c + 1
        lr, rr = self._excel_pair(v_row)
        st = self.sheet_diff.diffs.get((v_row, col), DiffStatus.EQUAL)
        lv = self.ws_left.cell(lr, col).value if lr is not None and self.ws_left is not None else None
        rv = self.ws_right.cell(rr, col).value if rr is not None and self.ws_right is not None else None
        self.selection_changed.emit(v_row, col, st.value, lv, rv)

    # ---------- Visual <-> Excel mapping ----------

    def _excel_pair(self, v_row: int) -> tuple[int | None, int | None]:
        return self.sheet_diff.excel_rows_at(v_row)

    # ---------- Inline edits ----------

    def _on_left_edited(self, item: QTableWidgetItem) -> None:
        if self._programmatic or self.ws_left is None:
            return
        v_row = item.row() + 1
        c = item.column() + 1
        lr, _ = self._excel_pair(v_row)
        if lr is None:
            return  # gap row — shouldn't be editable
        parsed = _parse_text(item.text())
        self.ws_left.cell(lr, c).value = parsed
        self._programmatic = True
        item.setData(Qt.ItemDataRole.UserRole, parsed)
        item.setText(_display_text(parsed))
        self._programmatic = False
        self._refresh_cell_status(v_row, c)
        self.left_modified.emit()

    def _on_right_edited(self, item: QTableWidgetItem) -> None:
        if self._programmatic or self.ws_right is None:
            return
        v_row = item.row() + 1
        c = item.column() + 1
        _, rr = self._excel_pair(v_row)
        if rr is None:
            return
        parsed = _parse_text(item.text())
        self.ws_right.cell(rr, c).value = parsed
        self._programmatic = True
        item.setData(Qt.ItemDataRole.UserRole, parsed)
        item.setText(_display_text(parsed))
        self._programmatic = False
        self._refresh_cell_status(v_row, c)
        self.right_modified.emit()

    # ---------- Copy actions (arrows) ----------

    def _current_cell(self) -> tuple[int, int] | None:
        r = self.left_table.currentRow()
        c = self.left_table.currentColumn()
        if r < 0 or c < 0:
            return None
        return r + 1, c + 1

    def can_copy(self) -> bool:
        return self.sheet_diff.presence == SheetPresence.BOTH and self._current_cell() is not None

    def copy_left_to_right(self) -> str | None:
        """Returns None on success, or an explanation string if the operation can't be done."""
        if self.ws_left is None or self.ws_right is None:
            return "No hay datos para copiar."
        pos = self._current_cell()
        if pos is None:
            return "Selecciona una celda primero."
        v_row, c = pos
        lr, rr = self._excel_pair(v_row)
        if lr is None:
            return "A no tiene fila en esta posición — nada que copiar."
        if rr is None:
            return "B no tiene fila aquí. Cambia a alineación posicional o crea la fila a mano."
        value = self.ws_left.cell(lr, c).value
        self._apply_programmatic(self.right_table, self.ws_right, v_row, c, rr, value)
        self.right_modified.emit()
        self._refresh_cell_status(v_row, c)
        return None

    def copy_right_to_left(self) -> str | None:
        if self.ws_left is None or self.ws_right is None:
            return "No hay datos para copiar."
        pos = self._current_cell()
        if pos is None:
            return "Selecciona una celda primero."
        v_row, c = pos
        lr, rr = self._excel_pair(v_row)
        if rr is None:
            return "B no tiene fila en esta posición — nada que copiar."
        if lr is None:
            return "A no tiene fila aquí. Cambia a alineación posicional o crea la fila a mano."
        value = self.ws_right.cell(rr, c).value
        self._apply_programmatic(self.left_table, self.ws_left, v_row, c, lr, value)
        self.left_modified.emit()
        self._refresh_cell_status(v_row, c)
        return None

    def _apply_programmatic(self, table: QTableWidget, ws, v_row: int, c: int, excel_row: int, value: Any) -> None:
        ws.cell(excel_row, c).value = value
        self._programmatic = True
        table.blockSignals(True)
        item = table.item(v_row - 1, c - 1)
        if item is None:
            item = QTableWidgetItem("")
            table.setItem(v_row - 1, c - 1, item)
        item.setText(_display_text(value))
        item.setData(Qt.ItemDataRole.UserRole, value)
        text = _display_text(value)
        item.setToolTip(text if len(text) > 20 else "")
        table.blockSignals(False)
        self._programmatic = False

    def _refresh_cell_status(self, v_row: int, c: int) -> None:
        from src.core.diff_engine import cell_status  # local import to keep top tidy
        lr, rr = self._excel_pair(v_row)
        lv = self.ws_left.cell(lr, c).value if lr is not None and self.ws_left is not None else None
        rv = self.ws_right.cell(rr, c).value if rr is not None and self.ws_right is not None else None
        new_status = cell_status(lv, rv)
        key = (v_row, c)
        if new_status == DiffStatus.EQUAL:
            self.sheet_diff.diffs.pop(key, None)
        else:
            self.sheet_diff.diffs[key] = new_status
        self._programmatic = True
        self.left_table.blockSignals(True)
        self.right_table.blockSignals(True)
        self._paint(self.left_table, v_row, c, cell_color(new_status.value, "left"))
        self._paint(self.right_table, v_row, c, cell_color(new_status.value, "right"))
        self.left_table.blockSignals(False)
        self.right_table.blockSignals(False)
        self._programmatic = False
        self._emit_selection(v_row - 1, c - 1)

    # ---------- Diff navigation ----------

    def goto_next_diff(self) -> bool:
        keys = sorted(self.sheet_diff.diffs.keys())
        if not keys:
            return False
        cur = self._current_cell()
        if cur is None:
            target = keys[0]
        else:
            target = next((k for k in keys if k > cur), keys[0])
        self._goto(*target)
        return True

    def goto_prev_diff(self) -> bool:
        keys = sorted(self.sheet_diff.diffs.keys())
        if not keys:
            return False
        cur = self._current_cell()
        if cur is None:
            target = keys[-1]
        else:
            target = next((k for k in reversed(keys) if k < cur), keys[-1])
        self._goto(*target)
        return True

    def _goto(self, r: int, c: int) -> None:
        self.left_table.setCurrentCell(r - 1, c - 1)

    # ---------- Shortcuts ----------

    def _install_shortcuts(self) -> None:
        sc_right = QShortcut(QKeySequence("Alt+Right"), self)
        sc_right.activated.connect(self.copy_left_to_right)
        sc_left = QShortcut(QKeySequence("Alt+Left"), self)
        sc_left.activated.connect(self.copy_right_to_left)
        sc_next = QShortcut(QKeySequence("F8"), self)
        sc_next.activated.connect(self.goto_next_diff)
        sc_prev = QShortcut(QKeySequence("Shift+F8"), self)
        sc_prev.activated.connect(self.goto_prev_diff)
