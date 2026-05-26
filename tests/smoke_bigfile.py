"""Stress test: build two Excels with inflated dimensions and verify the app doesn't choke."""
from __future__ import annotations

import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import PatternFill
from PySide6.QtCore import QEventLoop, QTimer
from PySide6.QtWidgets import QApplication


def build_inflated(path: Path, modify: bool) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    # Real content in first 100 rows × 10 cols
    for r in range(1, 101):
        for c in range(1, 11):
            value = r * 100 + c
            if modify and r == 50 and c == 5:
                value = 99999  # one diff
            ws.cell(r, c).value = value
    # Inflate phantom rows: write a styled-but-empty cell waaay down
    # This is what makes openpyxl report a huge max_row
    fill = PatternFill("solid", fgColor="FFFFFF")
    ws.cell(50000, 50).fill = fill
    wb.save(path)


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="cmp_big_"))
    a = tmp / "a.xlsx"
    b = tmp / "b.xlsx"
    print("Building inflated workbooks...")
    t0 = time.time()
    build_inflated(a, modify=False)
    build_inflated(b, modify=True)
    print(f"  built in {time.time()-t0:.1f}s")

    # Test the engine
    from src.core.diff_engine import compute_workbook_diff
    print("Loading workbooks...")
    t0 = time.time()
    wb_l = openpyxl.load_workbook(a)
    wb_r = openpyxl.load_workbook(b)
    print(f"  loaded in {time.time()-t0:.1f}s")
    print(f"  openpyxl reports max_row={wb_l['Datos'].max_row}, max_col={wb_l['Datos'].max_column}")
    print("Computing diff...")
    t0 = time.time()
    diff = compute_workbook_diff(wb_l, wb_r, str(a), str(b))
    print(f"  diff in {time.time()-t0:.1f}s, total diffs={diff.total_diffs}")
    sd = diff.sheets["Datos"]
    print(f"  Datos: max_row={sd.max_row}, max_col={sd.max_col}, diffs={sd.diff_count}")
    assert sd.diff_count == 1
    assert sd.max_row == 100, f"max_row should be content-based (100), got {sd.max_row}"

    # Test the UI panel materialization
    app = QApplication.instance() or QApplication(sys.argv)
    from src.ui.main_window import MainWindow, TabPanel
    w = MainWindow()
    w.show()
    w.left_path.setText(str(a))
    w.right_path.setText(str(b))
    w._start_compare()
    loop = QEventLoop()
    def done():
        if w.diff is not None:
            loop.quit()
        else:
            QTimer.singleShot(50, done)
    QTimer.singleShot(50, done)
    loop.exec()

    panel = w.tabs.widget(0)
    assert isinstance(panel, TabPanel)
    assert panel.view is not None
    print(f"  UI materialized OK")
    print("[OK] Big-file safety test passed.")
    return 0


if __name__ == "__main__":
    import traceback
    try:
        sys.exit(main())
    except Exception:
        traceback.print_exc()
        sys.exit(1)
