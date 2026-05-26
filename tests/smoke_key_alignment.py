"""Smoke test for key-based row alignment."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from PySide6.QtWidgets import QApplication

from src.core.diff_engine import (
    AlignmentMode,
    DiffStatus,
    compute_workbook_diff,
    detect_key_column,
    realign_sheet,
)
from src.ui.diff_view import DiffView


def main() -> int:
    app = QApplication.instance() or QApplication(sys.argv)

    left_path = str(ROOT / "ejemplos" / "antiguo.xlsx")
    right_path = str(ROOT / "ejemplos" / "nuevo.xlsx")
    wb_l = openpyxl.load_workbook(left_path, data_only=False)
    wb_r = openpyxl.load_workbook(right_path, data_only=False)
    diff = compute_workbook_diff(wb_l, wb_r, left_path, right_path)
    sd = diff.sheets["Productos"]
    print(f"POSITIONAL: max_row={sd.max_row}, diffs={sd.diff_count}, alignment={sd.alignment_mode.value}")
    assert sd.alignment_mode == AlignmentMode.POSITIONAL
    assert sd.diff_count == 31

    # Detect key column
    ws_l = wb_l["Productos"]
    ws_r = wb_r["Productos"]
    key = detect_key_column(ws_l, sd.src_max_row_left, sd.src_max_col_left)
    print(f"Detected key column: {key}  (1=SKU)")
    assert key == 1

    # Re-align using SKU
    realign_sheet(sd, ws_l, ws_r, AlignmentMode.KEY_BASED, key_col=1)
    print(f"KEY_BASED: max_row={sd.max_row}, diffs={sd.diff_count}, key_col={sd.key_column}")
    assert sd.alignment_mode == AlignmentMode.KEY_BASED
    assert sd.key_column == 1

    # Walk aligned rows: P005 onwards should be paired by SKU, not by Excel row
    # In old: row 6 = P005 Lechuga; In new: row 5 = P005 Lechuga.
    # So one of the aligned visual rows should be (6, 5).
    pairs = sd.aligned_rows
    assert (6, 5) in pairs, f"Expected (6,5) pairing for P005; got pairs {pairs}"
    assert (5, None) in pairs, "P004 (Tomate) should appear as left-only"
    assert (None, 9) in pairs or (None, 10) in pairs, "New products should be right-only at the bottom"
    print("[OK] Aligned rows look correct.")

    # Create the view and confirm it renders
    view = DiffView(ws_l, ws_r, sd, left_name="antiguo.xlsx", right_name="nuevo.xlsx")
    assert view.left_table.rowCount() == len(sd.aligned_rows)
    assert view.right_table.rowCount() == len(sd.aligned_rows)
    print(f"[OK] DiffView renders {view.left_table.rowCount()} visual rows.")

    # Verify a gap row (left-only) has the left side editable and right side non-editable
    for v_idx, (lr, rr) in enumerate(sd.aligned_rows, start=1):
        if lr is not None and rr is None:
            item_l = view.left_table.item(v_idx - 1, 0)
            item_r = view.right_table.item(v_idx - 1, 0)
            from PySide6.QtCore import Qt
            assert item_l.flags() & Qt.ItemFlag.ItemIsEditable, "left side of left-only row should be editable"
            assert not (item_r.flags() & Qt.ItemFlag.ItemIsEditable), "right side of left-only row should NOT be editable"
            print(f"[OK] Visual row {v_idx} (left-only): editability locked correctly.")
            break

    # Copy attempt on a gap row should fail gracefully (return message instead of crashing)
    for v_idx, (lr, rr) in enumerate(sd.aligned_rows, start=1):
        if lr is not None and rr is None:
            view.left_table.setCurrentCell(v_idx - 1, 0)
            msg = view.copy_left_to_right()
            assert msg is not None, "copy_left_to_right on gap should return a message"
            print(f"[OK] Copy on gap row returns: {msg!r}")
            break

    # Copy on a real paired row should succeed
    for v_idx, (lr, rr) in enumerate(sd.aligned_rows, start=1):
        if lr is not None and rr is not None and (v_idx, 5) in sd.diffs:
            # column E (stock) — let's pick this paired diff
            view.left_table.setCurrentCell(v_idx - 1, 4)
            before = ws_l.cell(lr, 5).value
            msg = view.copy_left_to_right()
            assert msg is None, f"Copy should succeed but got: {msg}"
            after = ws_r.cell(rr, 5).value
            assert after == before, f"Right E should now equal left E: {after} vs {before}"
            print(f"[OK] Copy left->right on paired visual row {v_idx} col E succeeded.")
            break

    print("\n[OK] Key-based alignment smoke test passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
