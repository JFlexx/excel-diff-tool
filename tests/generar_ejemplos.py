"""Genera dos Excel de ejemplo en ./ejemplos/ con cambios variados para probar el comparador."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


OUT_DIR = Path(__file__).resolve().parent.parent / "ejemplos"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _set_widths(ws, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -------------------- ANTIGUO --------------------

def build_antiguo(path: Path) -> None:
    wb = openpyxl.Workbook()

    # Hoja 1: Productos
    ws = wb.active
    ws.title = "Productos"
    ws.append(["SKU", "Producto", "Categoría", "Precio (€)", "Stock", "Total (€)"])
    _style_header(ws, 1, 6)
    _set_widths(ws, [10, 22, 16, 12, 8, 12])

    productos = [
        ("P001", "Manzana Golden",  "Fruta",       1.20, 150),
        ("P002", "Pera Conferencia","Fruta",       1.80, 90),
        ("P003", "Naranja Navel",   "Fruta",       1.50, 120),
        ("P004", "Tomate Rama",     "Verdura",     2.40, 60),
        ("P005", "Lechuga Iceberg", "Verdura",     1.10, 40),
        ("P006", "Pan Integral",    "Panadería",   2.20, 25),
        ("P007", "Leche Entera",    "Lácteos",     1.05, 200),
        ("P008", "Queso Manchego",  "Lácteos",    12.50, 18),
    ]
    for i, (sku, nombre, cat, precio, stock) in enumerate(productos, start=2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=nombre)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=precio).number_format = "#,##0.00 €"
        ws.cell(row=i, column=5, value=stock)
        ws.cell(row=i, column=6, value=f"=D{i}*E{i}").number_format = "#,##0.00 €"

    # Hoja 2: Clientes
    ws2 = wb.create_sheet("Clientes")
    ws2.append(["ID", "Nombre", "Ciudad", "Email", "Activo"])
    _style_header(ws2, 1, 5)
    _set_widths(ws2, [8, 24, 14, 28, 8])
    clientes = [
        (1, "Industrias Pérez S.L.",   "Madrid",    "pedidos@perez.es",      "Sí"),
        (2, "Distribuciones García",   "Barcelona", "compras@garcia.com",    "Sí"),
        (3, "Logística Norte",         "Bilbao",    "info@lognorte.es",      "Sí"),
        (4, "Comercial Sur",           "Sevilla",   "ventas@comsur.com",     "No"),
        (5, "Almacenes Centro",        "Valladolid","centro@almacen.es",     "Sí"),
    ]
    for c in clientes:
        ws2.append(c)

    # Hoja 3: Resumen (con fórmulas que referencian Productos)
    ws3 = wb.create_sheet("Resumen")
    ws3["A1"] = "Indicador"
    ws3["B1"] = "Valor"
    _style_header(ws3, 1, 2)
    _set_widths(ws3, [28, 16])
    ws3["A2"] = "Nº de productos"
    ws3["B2"] = "=COUNTA(Productos!A2:A100)"
    ws3["A3"] = "Stock total"
    ws3["B3"] = "=SUM(Productos!E2:E100)"
    ws3["A4"] = "Valor inventario (€)"
    ws3["B4"] = "=SUM(Productos!F2:F100)"
    ws3["B4"].number_format = "#,##0.00 €"

    # Hoja 4: solo en el antiguo (se eliminó en el nuevo)
    ws4 = wb.create_sheet("Proveedores_OLD")
    ws4.append(["Proveedor", "Contacto"])
    _style_header(ws4, 1, 2)
    ws4.append(["Frutas del Valle", "valle@frutas.com"])
    ws4.append(["Lácteos Norte",    "norte@lacteos.com"])

    wb.save(path)


# -------------------- NUEVO --------------------

def build_nuevo(path: Path) -> None:
    wb = openpyxl.Workbook()

    # Hoja 1: Productos (con cambios)
    ws = wb.active
    ws.title = "Productos"
    ws.append(["SKU", "Producto", "Categoría", "Precio (€)", "Stock", "Total (€)"])
    _style_header(ws, 1, 6)
    _set_widths(ws, [10, 22, 16, 12, 8, 12])

    productos = [
        ("P001", "Manzana Golden",     "Fruta",       1.35, 180),    # cambia precio + stock
        ("P002", "Pera Conferencia",   "Fruta",       1.80, 90),     # igual
        ("P003", "Naranja Navel",      "Fruta",       1.60, 120),    # cambia precio
        # P004 (Tomate Rama) eliminado
        ("P005", "Lechuga Iceberg",    "Verdura",     1.10, 55),     # cambia stock
        ("P006", "Pan Integral",       "Panadería",   2.30, 30),     # cambia precio + stock
        ("P007", "Leche Entera",       "Lácteos",     1.05, 200),    # igual
        ("P008", "Queso Manchego DOP", "Lácteos",    13.90, 20),     # cambia nombre, precio, stock
        ("P009", "Yogur Natural",      "Lácteos",     0.75, 300),    # nuevo
        ("P010", "Aceite Oliva Virgen","Despensa",    9.20, 45),     # nuevo
    ]
    for i, (sku, nombre, cat, precio, stock) in enumerate(productos, start=2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=nombre)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=precio).number_format = "#,##0.00 €"
        ws.cell(row=i, column=5, value=stock)
        ws.cell(row=i, column=6, value=f"=D{i}*E{i}").number_format = "#,##0.00 €"

    # Hoja 2: Clientes (con cambios sutiles)
    ws2 = wb.create_sheet("Clientes")
    ws2.append(["ID", "Nombre", "Ciudad", "Email", "Activo"])
    _style_header(ws2, 1, 5)
    _set_widths(ws2, [8, 24, 14, 28, 8])
    clientes = [
        (1, "Industrias Pérez S.L.",      "Madrid",    "pedidos@perez.es",        "Sí"),    # igual
        (2, "Distribuciones García S.A.", "Barcelona", "compras@garcia.com",      "Sí"),    # cambia razón social
        (3, "Logística Norte",            "Vitoria",   "info@lognorte.es",        "Sí"),    # cambia ciudad
        (4, "Comercial Sur",              "Sevilla",   "ventas@comsur.com",       "Sí"),    # cambia activo No->Sí
        (5, "Almacenes Centro",           "Valladolid","centro@almacenes.es",     "Sí"),    # cambia email
        (6, "Mercados Levante",           "Valencia",  "info@mercadolevante.es",  "Sí"),    # nuevo
    ]
    for c in clientes:
        ws2.append(c)

    # Hoja 3: Resumen (con una fórmula cambiada y un indicador nuevo)
    ws3 = wb.create_sheet("Resumen")
    ws3["A1"] = "Indicador"
    ws3["B1"] = "Valor"
    _style_header(ws3, 1, 2)
    _set_widths(ws3, [28, 16])
    ws3["A2"] = "Nº de productos"
    ws3["B2"] = "=COUNTA(Productos!A2:A200)"   # cambia el rango
    ws3["A3"] = "Stock total"
    ws3["B3"] = "=SUM(Productos!E2:E200)"
    ws3["A4"] = "Valor inventario (€)"
    ws3["B4"] = "=SUM(Productos!F2:F200)"
    ws3["B4"].number_format = "#,##0.00 €"
    ws3["A5"] = "Precio medio (€)"             # indicador nuevo
    ws3["B5"] = "=AVERAGE(Productos!D2:D200)"
    ws3["B5"].number_format = "#,##0.00 €"

    # Hoja 4: solo en el nuevo (añadida)
    ws4 = wb.create_sheet("Promociones")
    ws4.append(["SKU", "Descuento (%)", "Fecha fin"])
    _style_header(ws4, 1, 3)
    ws4.append(["P001", 10, "2026-06-30"])
    ws4.append(["P008", 15, "2026-06-15"])
    ws4.append(["P010", 20, "2026-07-31"])

    wb.save(path)


def main() -> int:
    OUT_DIR.mkdir(exist_ok=True)
    antiguo = OUT_DIR / "antiguo.xlsx"
    nuevo = OUT_DIR / "nuevo.xlsx"
    build_antiguo(antiguo)
    build_nuevo(nuevo)
    print(f"Generados:\n  {antiguo}\n  {nuevo}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
